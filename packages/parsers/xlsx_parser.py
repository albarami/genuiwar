"""XLSX parser — extracts sheet data with header detection."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from packages.parsers.base import BaseParser, ParseError, ParseResult
from packages.parsers.registry import register_parser
from packages.schemas.document import FileDocument
from packages.schemas.enums import FileType
from packages.schemas.evidence import CitationAnchor, EvidenceChunk

_CHUNK_ROW_SIZE = 50


@register_parser(FileType.XLSX)
class XlsxParser(BaseParser):
    """Parse .xlsx files into evidence chunks with sheet + row-range citation anchors."""

    def parse(self, file_path: Path, file_doc: FileDocument) -> ParseResult:
        """Parse an XLSX file into evidence chunks."""
        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
        except (InvalidFileException, Exception) as exc:
            raise ParseError(str(file_path), f"Cannot open XLSX: {exc}") from exc

        chunks: list[EvidenceChunk] = []
        warnings: list[str] = []
        sheet_names: list[str] = wb.sheetnames

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])

            if not rows:
                warnings.append(f"Sheet '{sheet_name}' is empty")
                continue

            headers = rows[0] if rows else []

            for start in range(0, len(rows), _CHUNK_ROW_SIZE):
                batch = rows[start : start + _CHUNK_ROW_SIZE]
                lines = [" | ".join(r) for r in batch]
                end = min(start + _CHUNK_ROW_SIZE, len(rows))

                chunks.append(
                    EvidenceChunk(
                        file_id=file_doc.file_id,
                        content="\n".join(lines),
                        content_type="table",
                        citation_anchor=CitationAnchor(
                            file_id=file_doc.file_id,
                            sheet_name=sheet_name,
                            row_range=(start + 1, end),
                        ),
                        metadata={
                            "headers": " | ".join(headers),
                            "sheet": sheet_name,
                        },
                    )
                )

        wb.close()

        if not chunks:
            warnings.append("XLSX file produced no content chunks")

        return ParseResult(
            chunks=chunks,
            metadata={"sheet_names": sheet_names, "sheet_count": len(sheet_names)},
            warnings=warnings,
        )
